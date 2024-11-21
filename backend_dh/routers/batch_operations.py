from fastapi import APIRouter, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook
import httpx
import os
import asyncio  # 비동기 라이브러리 임포트
from pathlib import Path
from fastapi.templating import Jinja2Templates

router = APIRouter()
templates = Jinja2Templates(directory="C:/project/backend_dh/templates")  # 템플릿 파일 경로 설정

# 파일 업로드 HTML 페이지 반환
@router.get("/batch-operations", response_class=HTMLResponse)
async def get_batch_operations_page(request: Request):
    return templates.TemplateResponse("batch_operations.html", {"request": request})

# 이미지 일괄 다운로드
@router.post("/batch-operations")
async def batch_download_images(file: UploadFile = File(...)):
    # 1. 업로드 파일 저장 경로 설정
    upload_path = Path(file.filename).resolve()
    with open(upload_path, "wb") as f:
        f.write(await file.read())
    
    # 2. 엑셀 파일 파싱
    wb = load_workbook(filename=upload_path, data_only=True)
    sheet = wb.active
    tasks = []
    
    # 3. 다운로드 파일 저장 경로 설정
    download_path = upload_path.parent / "downloaded_images"
    os.makedirs(download_path, exist_ok=True)
    
    # 4. URL 다운로드 준비
    async with httpx.AsyncClient() as client:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 줄은 헤더일 경우
            post_id, img_url = row
            if img_url:  # img_url이 있는 경우에만
                tasks.append(download_image(client, img_url, download_path / f"{post_id}.jpg"))
                
        # 5. 비동기 다운로드
        await asyncio.gather(*tasks)
    
    return {"message": "이미지 다운로드 완료", "download_path": str(download_path)}

async def download_image(client, url, filename):
    response = await client.get(url)
    with open(filename, "wb") as f:
        f.write(response.content)
